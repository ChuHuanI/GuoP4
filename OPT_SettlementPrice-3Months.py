import requests
import pandas as pd
from io import StringIO
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

# 定義目標 URL
url = 'https://www.taifex.com.tw/cht/5/optIndxFSP'

# 爬取數據
response = requests.get(url)
response.encoding = 'utf-8'

# 解析 HTML 表格
tables = pd.read_html(StringIO(response.text))
df = tables[0]

# 只保留 "契約月份" 與 "臺指選擇權（TXO）" 欄位
df = df.iloc[:, [1, 2]]
df.columns = ["商品", "結算價"]

# 移除電子與金融選擇權的數據
df = df.dropna()

# 只保留三個月內的數據
df = df.head(12)

# 反轉數據，使其順序與 Excel 一致
df = df[::-1].reset_index(drop=True)

# 輸出 Excel（不計算價差，讓 Excel 內使用公式）
excel_filename = "結算價與價差分析.xlsx"
df.to_excel(excel_filename, index=False, sheet_name="數據")

# 載入 Excel
wb = load_workbook(excel_filename)
ws = wb["數據"]

# 設定 A 欄寬度為 9.88
ws.column_dimensions['A'].width = 9.88

# 設定 C2 為 0
ws["C2"].value = 0

# 從 C3 開始填入 Excel 公式：C3 = B3 - B2，C4 = B4 - B3，以此類推
for row in range(3, len(df) + 2):  # 從 Excel 第 3 列開始（對應 Python 的 index 2）
    ws[f"C{row}"].value = f"=B{row}-B{row-1}"

# 設定 X 軸標籤（商品名稱）
categories = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)

# 設定 Y 軸數據（結算價）
values = Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1)

# **📈 結算價折線圖**
line_chart = LineChart()
line_chart.title = "結算價走勢"
line_chart.y_axis.title = "結算價"
line_chart.x_axis.title = "商品"
line_chart.width = 16.5
line_chart.height = 8.6
line_chart.add_data(values, titles_from_data=True)
line_chart.set_categories(categories)
ws.add_chart(line_chart, "D1")

# **📊 價差直方圖**
bar_chart = BarChart()
bar_chart.title = "價差分佈"
bar_chart.y_axis.title = "價差"
bar_chart.x_axis.title = "商品"
bar_chart.width = 16.5
bar_chart.height = 8.6
bar_values = Reference(ws, min_col=3, min_row=1, max_row=len(df) + 1)
bar_chart.add_data(bar_values, titles_from_data=True)
bar_chart.set_categories(categories)
ws.add_chart(bar_chart, "O1")

# 儲存 Excel
wb.save(excel_filename)

print(f"Excel 檔案已儲存為 {excel_filename}，價差欄位已使用 Excel 公式！")
